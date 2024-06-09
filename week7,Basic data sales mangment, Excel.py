import openpyxl
import matplotlib.pyplot as plt 

def creat_excel_file(filename):
  #creat a new excel file and add in some data
  workbook = openpyxl.Workbook()
  sheet = workbook.active
  sheet.title="Data"
  
  #add headers
  headers = ["Months","Sales"]
  sheet.append(headers)
  
  #add sample data
  data = [
        ["January", 150],
        ["February",200],
        ["March",250],
        ["April",300],
        ["May",350],
        ["June",400],
        ["July",450],
        ["August",500],
        ["September",550],
        ["October",600],
        ["November",650],
        ["December",700]
    ]
 
  for row in data:
    sheet.append(row)
      
  workbook.save(filename) 
  print(f"The excel file {filename} was created successfully!")
   
def read_excel_file(filename):
  workbook =openpyxl.load_workbook(filename)
  sheet = workbook.active
   
  data = []
  for row in sheet.iter_rows(values_only = True):
     data.append(row)
    
  return data
  
def visualise_data(data):
  #Extreact month and sales data
  months = [row[0]for row in data [1:]] #Exclude header
  sales = [row[1]for row in data[1:]]  #Exclude header
  
  plt.figure(figsize=(10,6))
  plt.plot(months, sales, marker ='o', linestyle ='-',color ='b')
  plt.title('Monthly Sales Data')
  plt.xlabel('Month')
  plt.ylabel('Sales')
  plt.grid(True)
  plt.show()
  
def main():
  filename = 'sales_data.xlsx'
  
  #Creat Excel file
  creat_excel_file(filename)
  
  #Read data from Excelfile
  data =read_excel_file(filename)
  print("Data from excel file:")
  for row in data:
    print(row)
 #Visualize the data in a line paragraph   
  visualise_data(data)  
    
if __name__== "__main__":
  main()
  
  
  